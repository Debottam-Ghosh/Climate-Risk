"""
run_forecast.py
===============
Main script for Kalman-based rainfall and temperature forecasting.

Usage
-----
Set DISTRICT, DATA_FILE, and TRAIN_END in the USER PARAMETERS block, then run:
    python run_forecast.py

Plots produced
--------------
  Plot 1.1 — Rainfall: all simulated ensemble paths
  Plot 1.2 — Rainfall: observed vs forecast (mean, p05/p10/p90/p95, LTA, metrics)
  Plot 2.1 — Temperature: all simulated ensemble paths
  Plot 2.2 — Temperature: observed vs forecast (mean, p05/p10/p90/p95, LTA, metrics)

Data format expected
--------------------
Excel file with columns: district, Year, Month, rain_mean, temp_mean
"""

import numpy as np
import openpyxl
import matplotlib.pyplot as plt

from rainfall_model2    import run_rainfall_kalman, MONTHS
from temperature_model import run_temperature_kalman, MONTH_NAMES
from enso_helper       import load_and_align as load_enso
from drought_model     import run_drought_model, plot_drought

# ─────────────────────────────────────────────────────────────────────────────
# ★  USER PARAMETERS  ← only change this block
# ─────────────────────────────────────────────────────────────────────────────

DISTRICT  = 'Bankura'

#FILE_LOC  = '/Users/dipankarmondal/Library/CloudStorage/OneDrive-IndianInstituteofTechnologyGuwahati/2026/Academic/Climate Challenge/Analysis/'
DATA_FILE = 'dicra_data_WB.xlsx'

# ── Optional: path to NOAA ONI CSV for ENSO covariate ───────────────────────
# Set to None to run without ENSO (default behaviour).
# Set to the file path to activate the ENSO teleconnection in the rain model.
#ONI_FILE  = None
ONI_FILE = 'ONI_text_to_csv_columns.csv'   # ← uncomment & edit

TRAIN_END = 2022   # last year used for training (test = everything after)
N_PATHS   = 5000   # Monte-Carlo paths
SEED      = int(np.random.uniform(1,100))


# Rainfall Kalman process noise
Q_MU0   = 1e-4
Q_GAMMA = 1e-7

# Temperature DLM state-evolution noise
DLM_Q = 0.05

# Context window shown left of the forecast boundary
CTX_MONTHS = 18

# ── Drought parameters ───────────────────────────────────────────────────────
SPI_SCALE = 3      # accumulation scale in months (1, 3, 6, or 12)
# LAT_DEG is auto-read from the dataset's latitude column.
# Override here only if your dataset lacks a latitude column:
# LAT_DEG = 17.0

# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_district(filepath, district):
    """
    Load year, month (0-indexed), rain_mean, temp_mean,
    latitude, longitude for one district.
    latitude/longitude columns are optional — returns NaN if absent.
    """
    wb   = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws   = wb.active
    hdrs = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def _col(*candidates):
        for name in candidates:
            if name in hdrs:
                return hdrs.index(name)
        return None   # optional columns return None instead of raising

    def _col_required(*candidates):
        idx = _col(*candidates)
        if idx is None:
            raise ValueError(f"Column not found. Tried: {candidates}. "
                             f"Available: {hdrs}")
        return idx

    i_d   = _col_required('district')
    i_y   = _col_required('Year', 'year')
    i_m   = _col_required('Month', 'month')
    i_r   = _col_required('rain_mean', 'Rain_mean', 'RAIN_MEAN')
    i_t   = _col_required('temp_mean', 'Temp_mean', 'TEMP_MEAN')
    i_lat = _col('latitude',  'Latitude',  'LAT',  'lat')
    i_lon = _col('longitude', 'Longitude', 'LON',  'lon')

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[i_d] != district:
            continue
        if row[i_r] is None or row[i_t] is None:
            continue
        lat = float(row[i_lat]) if (i_lat is not None and row[i_lat] is not None) else np.nan
        lon = float(row[i_lon]) if (i_lon is not None and row[i_lon] is not None) else np.nan
        rows.append((int(row[i_y]), int(row[i_m]) - 1,
                     float(row[i_r]), float(row[i_t]),
                     lat, lon))
    wb.close()

    if not rows:
        raise ValueError(
            f"District '{district}' not found or has no complete rows in "
            f"{filepath}.\nCheck the exact spelling.")

    rows.sort()
    yrs  = np.array([r[0] for r in rows])
    mos  = np.array([r[1] for r in rows])
    rain = np.array([r[2] for r in rows])
    temp = np.array([r[3] for r in rows])
    # lat/lon are constant per district — take the first non-NaN value
    lats = np.array([r[4] for r in rows])
    lons = np.array([r[5] for r in rows])
    lat  = float(lats[~np.isnan(lats)][0]) if (~np.isnan(lats)).any() else np.nan
    lon  = float(lons[~np.isnan(lons)][0]) if (~np.isnan(lons)).any() else np.nan
    return yrs, mos, rain, temp, lat, lon


# ─────────────────────────────────────────────────────────────────────────────
# LOAD & VALIDATE
# ─────────────────────────────────────────────────────────────────────────────

print("=" * 60)
print(f"  Kalman Forecast  |  {DISTRICT}")
print("=" * 60)
print("  Loading data ...")

yrs, mos, rain, temp, LAT_DEG, LON_DEG = load_district(DATA_FILE, DISTRICT)
train_mask = yrs <= TRAIN_END
T_tr_total = int(train_mask.sum())
T_te_total = int((~train_mask).sum())

if T_tr_total < 24:
    raise RuntimeError(f"Only {T_tr_total} training months — need ≥ 24.")
if T_te_total < 3:
    raise RuntimeError(f"Only {T_te_total} test months — need ≥ 3.")

print(f"  {len(yrs)} months  |  "
      f"Train: {T_tr_total} "
      f"({yrs[train_mask][0]}–{yrs[train_mask][-1]})  |  "
      f"Test: {T_te_total} "
      f"({yrs[~train_mask][0]}–{yrs[~train_mask][-1]})")

# Lat/lon auto-read from dataset
if not np.isnan(LAT_DEG):
    print(f"  Coordinates: lat={LAT_DEG:.4f}°N  lon={LON_DEG:.4f}°E "
          f"(auto-read from dataset)")
else:
    print("  WARNING: No latitude column found — defaulting LAT_DEG=20.0 "
          "for Thornthwaite PET. Set LAT_DEG manually in user parameters.")
    LAT_DEG = 20.0   # safe fallback for India

# ─────────────────────────────────────────────────────────────────────────────
# ENSO (optional)
# ─────────────────────────────────────────────────────────────────────────────

if ONI_FILE:
    print(f"\n── ENSO (ONI) ──────────────────────────────────────────")
    enso_aligned, _ = load_enso(ONI_FILE, yrs, mos, standardise=True)
else:
    enso_aligned = None   # rainfall_model will use zeros internally


# ─────────────────────────────────────────────────────────────────────────────
# RUN MODELS
# ─────────────────────────────────────────────────────────────────────────────

print("\n── Rainfall model ──────────────────────────────────────")
R = run_rainfall_kalman(
        rain, mos, yrs, train_mask,
        n_paths=N_PATHS, seed=SEED,
        q_mu0=Q_MU0, q_gamma=Q_GAMMA,
        enso=enso_aligned)

print("\n── Temperature model ───────────────────────────────────")
T = run_temperature_kalman(
        temp, mos, yrs, train_mask,
        n_paths=N_PATHS, seed=SEED, q=DLM_Q)


# ─────────────────────────────────────────────────────────────────────────────
# AXIS / LABEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _axis_labels(years_all, months_all, T_tr, T_te, ctx):
    yr_ctx = years_all[T_tr - ctx: T_tr]
    mo_ctx = months_all[T_tr - ctx: T_tr]
    yr_te  = years_all[T_tr:]
    mo_te  = months_all[T_tr:]

    ctx_lbl = [f"{yr_ctx[i]} {MONTH_NAMES[mo_ctx[i]]}" for i in range(ctx)]
    fc_lbl  = [f"{yr_te[i]}  {MONTH_NAMES[mo_te[i]]}"  for i in range(T_te)]

    x_ctx = np.arange(-ctx, 0)
    x_fc  = np.arange(T_te)
    xt    = list(range(-ctx, 0, 3)) + list(range(0, T_te, 3))
    xl    = ([ctx_lbl[i] for i in range(0, ctx, 3)] +
             [fc_lbl[i]  for i in range(0, T_te, 3)])
    return x_ctx, x_fc, xt, xl


def _metrics_box(ax, acc, unit):
    """Annotate metrics in the upper-left corner."""
    txt = (f"RMSE  = {acc['rmse']:.3f} {unit}\n"
           f"MAE   = {acc['mae']:.3f} {unit}\n"
           f"BIAS  = {acc['bias']:+.3f} {unit}\n"
           f"SKILL = {acc['skill']:.3f}\n"
           f"90% PI cov. = {acc['cov90']:.0f}%")
    ax.text(0.013, 0.975, txt,
            transform=ax.transAxes, fontsize=7.5,
            va='top', ha='left', family='monospace',
            bbox=dict(boxstyle='round,pad=0.4',
                      fc='white', ec='#bbbbbb', alpha=0.88))


# ─────────────────────────────────────────────────────────────────────────────
# COLOURS
# ─────────────────────────────────────────────────────────────────────────────

C_KF  = '#1D4E89'
C_TMP = '#7B2D8B'
C_OBS = '#D85A30'
C_LTA = '#2E8B57'
C_TR  = '#888780'


# ─────────────────────────────────────────────────────────────────────────────
# ══ PLOT 1.1  Rainfall — all simulated paths ═════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────

print("\n  Generating plots ...")

ctx_r = min(CTX_MONTHS, R['T_tr'])
T_te_r = R['T_te']
x_ctx_r, x_fc_r, xt_r, xl_r = _axis_labels(
    R['years_all'], R['months_all'], R['T_tr'], T_te_r, ctx_r)

lta_fc_r  = R['lta_mo'][R['mo_te']]
lta_ctx_r = R['lta_mo'][R['months_all'][R['T_tr'] - ctx_r: R['T_tr']]]

fig11, ax = plt.subplots(figsize=(14, 5))
fig11.suptitle(
    f"{DISTRICT}  —  Plot 1.1: Rainfall Ensemble Paths  "
    f"({N_PATHS} paths, Kalman Log-OU)",
    fontsize=11, fontweight='bold')

# Paths (capped at 500 for legibility)
n_show = min(500, N_PATHS)
idx_show = np.linspace(0, N_PATHS - 1, n_show, dtype=int)
for i in idx_show:
    ax.plot(x_fc_r, R['X_kalman'][i], lw=0.3, color=C_KF, alpha=0.06)

ax.plot(x_ctx_r, R['rain_tr'][-ctx_r:], lw=1.3, color=C_TR, ls='--',
        alpha=0.7, label='Training context')
ax.axvline(-0.5, lw=1.0, color='black', alpha=0.2, ls=':')
ax.plot(x_fc_r, R['fc']['mean'], lw=2.2, color=C_KF,
        zorder=5, label='Ensemble mean')
ax.scatter(x_fc_r, R['rain_te'], s=42, color=C_OBS, zorder=6,
           label='Observed', edgecolors='white', linewidths=0.5)
ax.plot(x_ctx_r, lta_ctx_r, lw=1.3, color=C_LTA, ls='-.',
        alpha=0.85, label='LTA (training)')
ax.plot(x_fc_r,  lta_fc_r,  lw=1.3, color=C_LTA, ls='-.', alpha=0.85)

ax.set_xticks(xt_r)
ax.set_xticklabels(xl_r, rotation=40, ha='right', fontsize=7)
ax.set_xlim(-ctx_r, T_te_r - 0.2); ax.set_ylim(bottom=0)
ax.set_ylabel('Rainfall (mm/day)', fontsize=10)
ax.legend(fontsize=8, ncol=4, loc='upper left')
ax.grid(True, alpha=0.18)
plt.tight_layout()


# ─────────────────────────────────────────────────────────────────────────────
# ══ PLOT 1.2  Rainfall — observed vs forecast ════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────

fig12, ax = plt.subplots(figsize=(14, 5))
fig12.suptitle(
    f"{DISTRICT}  —  Plot 1.2: Rainfall Forecast vs Observed  "
    f"(Kalman Log-OU)",
    fontsize=11, fontweight='bold')

# Shaded PI bands
ax.fill_between(x_fc_r, R['fc']['p05'], R['fc']['p95'],
                alpha=0.13, color=C_KF, label='5th–95th (90% PI)')
ax.fill_between(x_fc_r, R['fc']['p10'], R['fc']['p90'],
                alpha=0.22, color=C_KF, label='10th–90th (80% PI)')

# Quantile lines
for vals, ls_, pct in [(R['fc']['p95'], '--', '95th'),
                        (R['fc']['p90'], ':',  '90th'),
                        (R['fc']['p10'], ':',  '10th'),
                        (R['fc']['p05'], '--', '5th')]:
    ax.plot(x_fc_r, vals, lw=0.85, color=C_KF, ls=ls_, alpha=0.75)
    ax.annotate(pct, xy=(T_te_r - 1, vals[-1]),
                xytext=(T_te_r - 0.1, vals[-1]),
                fontsize=6.5, color=C_KF, va='center')

# Training context
ax.plot(x_ctx_r, R['rain_tr'][-ctx_r:], lw=1.2, color=C_TR,
        ls='--', alpha=0.6, label='Training context')
ax.axvline(-0.5, lw=1.0, color='black', alpha=0.2, ls=':')

# LTA
ax.plot(x_ctx_r, lta_ctx_r, lw=1.6, color=C_LTA, ls='-.', alpha=0.9,
        label='LTA (training climatology)')
ax.plot(x_fc_r,  lta_fc_r,  lw=1.6, color=C_LTA, ls='-.', alpha=0.9)

# Forecast mean
ax.plot(x_fc_r, R['fc']['mean'], lw=2.5, color=C_KF,
        zorder=5, label='Kalman mean')

# Observed — black edge when outside 90% PI
for i, (o, lo, hi) in enumerate(
        zip(R['rain_te'], R['fc']['p05'], R['fc']['p95'])):
    ec = 'black' if not (lo <= o <= hi) else 'none'
    ax.plot(i, o, 'o', ms=6, color=C_OBS,
            markeredgecolor=ec, markeredgewidth=1.6, zorder=7)
ax.plot([], [], 'o', color=C_OBS, label='Observed')
ax.plot([], [], 'o', color=C_OBS, markeredgecolor='black',
        markeredgewidth=1.6, label='Outside 90% PI')

_metrics_box(ax, R['acc'], 'mm/day')

ax.set_xticks(xt_r)
ax.set_xticklabels(xl_r, rotation=40, ha='right', fontsize=7)
ax.set_xlim(-ctx_r, T_te_r - 0.2); ax.set_ylim(bottom=0)
ax.set_ylabel('Rainfall (mm/day)', fontsize=10)
ax.legend(fontsize=8, ncol=3, loc='upper right')
ax.grid(True, alpha=0.18)
plt.tight_layout()


# ─────────────────────────────────────────────────────────────────────────────
# ══ PLOT 2.1  Temperature — all simulated paths ══════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────

ctx_t  = min(CTX_MONTHS, T['T_tr'])
T_te_t = T['T_te']
x_ctx_t, x_fc_t, xt_t, xl_t = _axis_labels(
    T['years_all'], T['months_all'], T['T_tr'], T_te_t, ctx_t)

lta_fc_t  = T['lta_mo'][T['mo_te']]
lta_ctx_t = T['lta_mo'][T['months_all'][T['T_tr'] - ctx_t: T['T_tr']]]

fig21, ax = plt.subplots(figsize=(14, 5))
fig21.suptitle(
    f"{DISTRICT}  —  Plot 2.1: Temperature Ensemble Paths  "
    f"({N_PATHS} paths, Kalman DLM-OU)",
    fontsize=11, fontweight='bold')

for i in idx_show:
    ax.plot(x_fc_t, T['paths'][i], lw=0.3, color=C_TMP, alpha=0.06)

ax.plot(x_ctx_t, T['temp_tr'][-ctx_t:], lw=1.3, color=C_TR,
        ls='--', alpha=0.7, label='Training context')
ax.axvline(-0.5, lw=1.0, color='black', alpha=0.2, ls=':')
ax.plot(x_fc_t, T['fc']['mean'], lw=2.2, color=C_TMP,
        zorder=5, label='Ensemble mean')
ax.scatter(x_fc_t, T['temp_te'], s=42, color=C_OBS, zorder=6,
           label='Observed', edgecolors='white', linewidths=0.5)
ax.plot(x_ctx_t, lta_ctx_t, lw=1.3, color=C_LTA, ls='-.',
        alpha=0.85, label='LTA (training)')
ax.plot(x_fc_t,  lta_fc_t,  lw=1.3, color=C_LTA, ls='-.', alpha=0.85)

ax.set_xticks(xt_t)
ax.set_xticklabels(xl_t, rotation=40, ha='right', fontsize=7)
ax.set_xlim(-ctx_t, T_te_t - 0.2)
ax.set_ylabel('Temperature (°C)', fontsize=10)
ax.legend(fontsize=8, ncol=4, loc='upper left')
ax.grid(True, alpha=0.18)
plt.tight_layout()


# ─────────────────────────────────────────────────────────────────────────────
# ══ PLOT 2.2  Temperature — observed vs forecast ══════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────

fig22, ax = plt.subplots(figsize=(14, 5))
fig22.suptitle(
    f"{DISTRICT}  —  Plot 2.2: Temperature Forecast vs Observed  "
    f"(Kalman DLM-OU)",
    fontsize=11, fontweight='bold')

ax.fill_between(x_fc_t, T['fc']['p05'], T['fc']['p95'],
                alpha=0.13, color=C_TMP, label='5th–95th (90% PI)')
ax.fill_between(x_fc_t, T['fc']['p10'], T['fc']['p90'],
                alpha=0.22, color=C_TMP, label='10th–90th (80% PI)')

for vals, ls_, pct in [(T['fc']['p95'], '--', '95th'),
                        (T['fc']['p90'], ':',  '90th'),
                        (T['fc']['p10'], ':',  '10th'),
                        (T['fc']['p05'], '--', '5th')]:
    ax.plot(x_fc_t, vals, lw=0.85, color=C_TMP, ls=ls_, alpha=0.75)
    ax.annotate(pct, xy=(T_te_t - 1, vals[-1]),
                xytext=(T_te_t - 0.1, vals[-1]),
                fontsize=6.5, color=C_TMP, va='center')

ax.plot(x_ctx_t, T['temp_tr'][-ctx_t:], lw=1.2, color=C_TR,
        ls='--', alpha=0.6, label='Training context')
ax.axvline(-0.5, lw=1.0, color='black', alpha=0.2, ls=':')

ax.plot(x_ctx_t, lta_ctx_t, lw=1.6, color=C_LTA, ls='-.', alpha=0.9,
        label='LTA (training climatology)')
ax.plot(x_fc_t,  lta_fc_t,  lw=1.6, color=C_LTA, ls='-.', alpha=0.9)

ax.plot(x_fc_t, T['fc']['mean'], lw=2.5, color=C_TMP,
        zorder=5, label='Kalman mean')

for i, (o, lo, hi) in enumerate(
        zip(T['temp_te'], T['fc']['p05'], T['fc']['p95'])):
    ec = 'black' if not (lo <= o <= hi) else 'none'
    ax.plot(i, o, '^', ms=6, color=C_OBS,
            markeredgecolor=ec, markeredgewidth=1.6, zorder=7)
ax.plot([], [], '^', color=C_OBS, label='Observed')
ax.plot([], [], '^', color=C_OBS, markeredgecolor='black',
        markeredgewidth=1.6, label='Outside 90% PI')

_metrics_box(ax, T['acc'], '°C')

ax.set_xticks(xt_t)
ax.set_xticklabels(xl_t, rotation=40, ha='right', fontsize=7)
ax.set_xlim(-ctx_t, T_te_t - 0.2)
ax.set_ylabel('Temperature (°C)', fontsize=10)
ax.legend(fontsize=8, ncol=3, loc='upper right')
ax.grid(True, alpha=0.18)
plt.tight_layout()


# ─────────────────────────────────────────────────────────────────────────────
# ══ DROUGHT MODEL — SPI & SPEI ════════════════════════════════════════════════
# ─────────────────────────────────────────────────────────────────────────────

print(f"\n── Drought model (SPI-{SPI_SCALE} & SPEI-{SPI_SCALE}) ──────────────")
DR = run_drought_model(R, T, scale=SPI_SCALE, lat_deg=LAT_DEG)

# Build context year/month arrays for the drought context window
yr_ctx_dr  = R['years_all'][R['T_tr'] - DR['ctx']: R['T_tr']]
mo_ctx_dr  = R['months_all'][R['T_tr'] - DR['ctx']: R['T_tr']]

fig3 = plot_drought(
    DR, DISTRICT,
    x_fc   = np.arange(R['T_te']),
    xt     = xt_r,
    xl     = xl_r,
    yr_te  = R['years_all'][R['T_tr']:],
    mo_te  = R['mo_te'],
    yr_ctx = yr_ctx_dr,
    mo_ctx = mo_ctx_dr,
)


# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

print(f"\n{'='*62}")
print(f"  {DISTRICT}  —  Forecast Summary")
print(f"{'='*62}")
print(f"  {'Metric':<24}  {'Rainfall':>14}  {'Temperature':>12}")
print(f"  {'-'*56}")
rows_summary = [
    ('RMSE',       'rmse',  'mm/day', '°C'),
    ('MAE',        'mae',   'mm/day', '°C'),
    ('BIAS',       'bias',  'mm/day', '°C'),
    ('SKILL',      'skill', '',       ''),
    ('90% PI cov', 'cov90', '%',      '%'),
]
for label, key, ur, ut in rows_summary:
    rv = R['acc'][key]; tv = T['acc'][key]
    fmt = '+.3f' if key == 'bias' else '.3f'
    print(f"  {label:<24}  {rv:{fmt}} {ur:<7}  {tv:{fmt}} {ut}")
print(f"\n  Rainfall  θ={R['sde']['theta']:.3f}  "
      f"AR(1)={R['ar1']:.3f}  γ={R['sde']['trend']:+.3f}/dec")
print(f"  Temp DLM  θ={T['dlm'].theta:.4f}  "
      f"half-life={T['dlm'].half_life():.2f} mo  "
      f"σ={T['dlm'].sigma:.4f}  q={DLM_Q}")
print(f"\n── Drought risk (mean over forecast horizon) ──────────")
for tag, prob_d in [('SPI', DR['spi_prob']), ('SPEI', DR['spei_prob'])]:
    print(f"  {tag}-{SPI_SCALE}  "
          f"P(moderate)={prob_d['moderate'].mean():.1f}%  "
          f"P(severe)={prob_d['severe'].mean():.1f}%  "
          f"P(extreme)={prob_d['extreme'].mean():.1f}%")
print(f"{'='*62}")

# ─────────────────────────────────────────────────────────────────────────────
# SAVE (uncomment to write PNGs)
# ─────────────────────────────────────────────────────────────────────────────
# tag = DISTRICT.replace(' ', '_')
# fig11.savefig(f'{tag}_plot1_1_rain_paths.png',    dpi=150, bbox_inches='tight')
# fig12.savefig(f'{tag}_plot1_2_rain_forecast.png', dpi=150, bbox_inches='tight')
# fig21.savefig(f'{tag}_plot2_1_temp_paths.png',    dpi=150, bbox_inches='tight')
# fig22.savefig(f'{tag}_plot2_2_temp_forecast.png', dpi=150, bbox_inches='tight')
# fig3.savefig(f'{tag}_plot3_drought_spi_spei.png', dpi=150, bbox_inches='tight')

plt.show()
print("Done.")
